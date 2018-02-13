#!python3
# 12multiplicationTable.py - Creates a multiplication table.

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if len(sys.argv) > 1:
    N = int(sys.argv[1])

    # Create bold, frozen pane headers in row 1 & column 1.
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.freeze_panes = 'B2'
    bold = Font(bold=True)
    for i in range(2, N+2):
        sheet['A' + str(i)].font = bold
        sheet['A' + str(i)].value = i-1
        sheet[get_column_letter(i) + '1'].font = bold
        sheet[get_column_letter(i) + '1'].value = i-1

    # Populate multiplication table
    for i in range(2, N+2):
        for j in range(2, N+2):
            sheet[get_column_letter(i) + str(j)] = (i-1)*(j-1)

    wb.save('multTable{}.xlsx'.format(N))

else:
    print('Command line call must include the dimension of the square multiplication table (call N)')
