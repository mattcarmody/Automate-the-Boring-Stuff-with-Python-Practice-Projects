import csv, openpyxl, os
from openpyxl.utils import get_column_letter

DIR = "/home/matt/AutomateBookLocal/spreadsheets/"

for _file in os.listdir(DIR):
    # Skip any non-Excel files.
    if _file.endswith(".xlsx"):
        wb = openpyxl.load_workbook("{}{}".format(DIR, _file))
        # Create a CSV file for each sheet
        for sheet_name in wb.get_sheet_names():
            csv_file = open("{}_{}.csv".format(_file[:-5], sheet_name), 'w', newline='')
            csv_writer = csv.writer(csv_file)
            sheet = wb.get_sheet_by_name(sheet_name)
            
            for row in range(1, sheet.max_row+1):
                row_data = []
                for column in range(1, sheet.max_column+1):
                    row_data.append(sheet["{}{}".format(get_column_letter(column), row)].value)
                csv_writer.writerow(row_data)
            csv_file.close()
            print("Saving {}".format(csv_file.name))
            
