#! /usr/bin/python3
# chap12PracProjCellInverter.py - Invert cells in a new spreadsheet.

import openpyxl
from openpyxl.utils import get_column_letter
import sys

if len(sys.argv) == 2:
    wb_old = openpyxl.load_workbook(sys.argv[1])
    sheet_old = wb_old.active

    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active

    for i in range(1, sheet_old.max_row + 1):
        for j in range(1, sheet_old.max_column + 1):
            sheet_new.cell(row=j, column=i).value = sheet_old.cell(row=i, column=j).value

    wb_new.save("invertedData.xlsx")
else:
    print("Command line call must include the name of the spreadsheet to be inverted (call filename)")
