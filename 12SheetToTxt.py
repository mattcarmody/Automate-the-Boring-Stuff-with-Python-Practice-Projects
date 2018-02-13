# Copy contents of spreadsheet to a txt file.

import openpyxl

files = ['rows2.txt', 'rowow2.txt', 'rowowow2.txt']

wb = openpyxl.load_workbook('txtToSheet.xlsx')
sheet = wb.active

for j in range(1, sheet.max_column+1):
	_file = open(files[j-1], 'w')
	for i in range(1, sheet.max_row+1):
		if sheet.cell(row=i, column=j).value:
			_file.write(sheet.cell(row=i, column=j).value)
	_file.close
