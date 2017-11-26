#!python3
# 12blankRowInserter.py - Insert M number of rows at row N in passed spreadsheet.

import sys, openpyxl
from openpyxl.utils import get_column_letter

# Store information passed from commmand line.
if len(sys.argv) > 1:
    N = int(sys.argv[1])   # when to start skipping rows
    M = int(sys.argv[2])   # how many rows to skip
    Z = sys.argv[3]        # file to read and play with
else:
    print('Something has gone horribly wrong!')

# Read in passed spreadsheet.
wb1 = openpyxl.load_workbook(Z)
sheet1 = wb1.active

# Write that in a new spreadsheet with a gap.
wb2 = openpyxl.Workbook()
sheet2 = wb2.active
for j in range(1, N):
    for i in range(1, sheet1.max_column+1):
        sheet2[get_column_letter(i) + str(j)] = sheet1[get_column_letter(i) + str(j)].value

for j in range(N, sheet1.max_row+1):
    for i in range(1, sheet1.max_column+1):
        sheet2[get_column_letter(i) + str(j+M)] = sheet1[get_column_letter(i) + str(j)].value

# Save file.
wb2.save('blankRow.xlsx')
