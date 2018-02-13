#!python3
# 12blankRowInserter.py - Insert M number of rows at row N in passed spreadsheet.

import sys, openpyxl
from openpyxl.utils import get_column_letter

# Store information passed from commmand line.
if len(sys.argv) >= 4:
    N = int(sys.argv[1])   # when to start skipping rows
    M = int(sys.argv[2])   # how many rows to skip
    wb_name = sys.argv[3]        # file to read and play with
else:
    print('Command line call should pass the row number to begin skipping at, how many rows to skip and the file name. (call N M file)')

# Read in passed spreadsheet.
wb_old = openpyxl.load_workbook(wb_name)
sheet_old = wb_old.active

# Write that in a new spreadsheet with a gap.
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active
for j in range(1, N):
    for i in range(1, sheet_old.max_column+1):
        sheet_new[get_column_letter(i) + str(j)] = sheet_old[get_column_letter(i) + str(j)].value

for j in range(N, sheet_old.max_row+1):
    for i in range(1, sheet_old.max_column+1):
        sheet_new[get_column_letter(i) + str(j+M)] = sheet_old[get_column_letter(i) + str(j)].value

wb_new.save('blankRow.xlsx')
